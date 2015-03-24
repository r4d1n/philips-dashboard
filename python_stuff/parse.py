import datetime
import pdb
from openpyxl import load_workbook

def format_date(date_string, time_string):
	print date_string.split('/')
	#print year, month, day
	#unix_time_day = datetime.datetime(year, month, day)


def main():
	wb = load_workbook(filename='pneumonia.xlsx')
	sheet1 = wb.get_sheet_by_name('Sheet1')
	data = {
		'Event': '[',
		'Temp': '[',
		'O2': '[',
		'WBC': '[',
		'BC': '[',
		'xray': '[',
		'therapy': '[',
	}
	relevant_columns = [(2, 'Event'), (3, 'Temp'), (4, 'O2'), (6, 'WBC'), (7, 'BC'), (9, 'xray'), (12, 'therapy')]
	row_number = 0
	for row in sheet1.iter_rows():
		if row_number > 6 and row[0].value and row[1].value:
			unix_time = row[0].value.strftime("%s")
			added_hour = int(str(row[1].value)[:-2])
			added_minute = int(str(row[1].value)[-2:])
			unix_time_total = int(unix_time) + added_minute*60 + added_hour*3600

			for col_number, label in relevant_columns:
				if row[col_number].value:
					string_to_add = '{ x:%s, y: %s} ,'%(unix_time_total, row[col_number].value)
					data[label] += string_to_add
					#data[label].append({"x":unix_time_total, "y":row[col_number].value})
		row_number += 1
	for key, value in data.iteritems():
		print key, value + ' ]'




			
		#print date
		row_number += 1

	

if __name__ == '__main__':
	
	main()